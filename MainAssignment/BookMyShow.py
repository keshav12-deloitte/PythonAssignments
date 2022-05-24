import sys
import openpyxl


def workbook(filename, sheetname):
    """this Function is used for accessing excel sheet by giving path of sheet and sheetname"""
    wb = openpyxl.load_workbook(filename)
    sh1 = wb[sheetname]
    row = sh1.max_row
    col = sh1.max_column
    return row, col, sh1, wb


def registeredUsersDict(self):
    """This Function is used for storing the already Registered users in a Dictionary and returning that Dictionary"""
    row, col, sh1, wb=workbook("RegisteredUserDetails.xlsx","Sheet1")
    userdetailsDict = {}
    for i in range(1, row + 2):
        cell_value_class = sh1.cell(i, 1).value
        cell_value_id = sh1.cell(i, 5).value
        userdetailsDict[cell_value_class] = cell_value_id
    return userdetailsDict


class AdminOfBookMyShow():

    def __init__(self, adminName, adminPassword):
        """This is a constructor for admin"""
        self.adminName = adminName
        self.adminPassword = adminPassword

    def adminOperations(self):
        """This Method is used for All Admin Operations."""
        while (True):
            print("******WelcomeAdmin*******\n1.Add New Movie Info\n2.Edit Movie\n3.Delete Movie\n4.Logout")
            choice = int(input("Enter the choice you want to perform: "))
            row, col, sh1, wb = workbook("MoviesInfo.xlsx", "Sheet1")
            if (choice == 4):
                break
            if (choice == 1):
                print("enter the details of the movie: ")
                row1 = 1
                column1 = 1
                for c in range(1, col + 1):
                    m = (sh1.cell(row1, column1).value)
                    sh1.cell(row + 1, c, value=input(f"Enter the {m}: "))
                    column1 += 1
                wb.save("MoviesInfo.xlsx")
                # break
            elif (choice == 2):
                movieNameToBeEdited = input("enter the movie name that you want to edit: ")
                for i in range(2, row + 1):
                    #print(row)
                    name = sh1.cell(i, column=1).value
                    if (movieNameToBeEdited == name):
                            for j in range(1, col + 1):
                                sh1.cell(i, j, value=input(f"Enter the {sh1.cell(1, j).value} new change: "))

                            # column1 += 1
                    else:
                        print(f"Their is no movie called : {movieNameToBeEdited} .please check again")
                wb.save("MoviesInfo.xlsx")
            elif (choice == 3):
                movieToBeDeleted = input("Enter the movie name you want to Delete: ")
                for i in range(2, row + 1):
                    name = sh1.cell(i, column=1).value
                    if (name == movieToBeDeleted):
                        sh1.delete_rows(i, True)
                wb.save("MoviesInfo.xlsx")